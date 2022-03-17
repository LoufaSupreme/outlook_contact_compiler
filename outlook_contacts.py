import openpyxl # must use version 2.6.2
import re

# open an excel workbook and return an openpyxl workbook instance
def open_wb(name):
    try:
        wb = openpyxl.load_workbook(name)
        print('Successfully loaded input worksheet')
        return wb
    except Exception as e:
        print(e)
        return

# format first and last name
def compile_name(full_name):
    first = full_name[0].strip('\'').strip(',') if len(full_name) > 0 and "@" not in full_name[0] else ""
    last = full_name[1].strip('\'').strip(',') if len(full_name) > 1 else ""
    return {'first': first, 'last': last}

# compares an excel sheet to another "master" excel sheet for duplicate email addresses
# if duplicates are found, removes those rows from the comparison wb and saves a new copy
def remove_duplicates(master_wb, comparison_wb):
    master_sheet = master_wb.active
    comparison_sheet = comparison_wb.active
    for i in range(comparison_sheet.max_row, 1, -1):
        comparison_email = comparison_sheet['A' + str(i)].value
        for j in range(2, master_sheet.max_row + 1):
            master_email = master_sheet['A' + str(j)].value
            if master_email == comparison_email:
                print(f'found match: {master_email} = {comparison_email}')
                comparison_sheet.delete_rows(i)
                break
    print('Removed duplicates')
    comparison_wb.save('compiled_noDuplicates.xlsx')

# takes a spreadsheet of a list of To, From and CC names/emails and copies unique/valid entries to a new output wb
# also pulls out company name from email domain name
def compile_contacts(input_wb, output_wb):    
    emails_found = []
    current_row_in_results = 2
    input_worksheet = input_wb.active
    output_worksheet = output_wb.active

    for i in range(2, input_worksheet.max_row + 1):
        try:
            from_email = input_worksheet['B' + str(i)].value.lower()
        except Exception as e:
            print(e, f'on line {i}')
            continue
        if from_email[0] != '/':
            try:
                emails_found.index(from_email)
            except:
                emails_found.append(from_email)

                full_name = input_worksheet['A' + str(i)].value.split(' ')
                first_name = compile_name(full_name)['first']
                last_name = compile_name(full_name)['last']

                try:
                    company = re.search('@(.*)\.[com|ca|org|net]', from_email).group(1)
                except:
                    company = ''

                output_worksheet['A' + str(current_row_in_results)].value = from_email
                output_worksheet['B' + str(current_row_in_results)].value = first_name
                output_worksheet['C' + str(current_row_in_results)].value = last_name
                output_worksheet['D' + str(current_row_in_results)].value = company
                
                current_row_in_results += 1

        to_emails = input_worksheet['D' + str(i)].value.split(';') if input_worksheet['D' + str(i)].value else []
        to_full_names = input_worksheet['C' + str(i)].value.split(';') if input_worksheet['C' + str(i)].value else []
        if len(to_emails) > 0:
            for j in range(0, len(to_emails)):
                if len(to_emails[j]) > 0 and to_emails[j][0] != '/':
                    to_emails[j] = to_emails[j].lower()
                    try:
                        emails_found.index(to_emails[j])
                    except:
                        emails_found.append(to_emails[j])

                        full_name = to_full_names[j].split(' ') if len(to_full_names) == len(to_emails) else []
                        first_name = compile_name(full_name)['first']
                        last_name = compile_name(full_name)['last']

                        try:
                            company = re.search('@(.*)\.[com|ca|org|net]', to_emails[j]).group(1)
                        except:
                            company = ''

                        output_worksheet['A' + str(current_row_in_results)].value = to_emails[j]
                        output_worksheet['B' + str(current_row_in_results)].value = first_name
                        output_worksheet['C' + str(current_row_in_results)].value = last_name
                        output_worksheet['D' + str(current_row_in_results)].value = company

                        current_row_in_results += 1

        cc_emails = input_worksheet['F' + str(i)].value.split(';') if input_worksheet['F' + str(i)].value else []
        cc_full_names = input_worksheet['E' + str(i)].value.split(';') if input_worksheet['E' + str(i)].value else []
        if len(cc_emails) > 0:
            for j in range(0, len(cc_emails)):
                if cc_emails[j][0] != '/':
                    cc_emails[j] = cc_emails[j].lower()
                    try:
                        emails_found.index(cc_emails[j])
                    except:
                        emails_found.append(cc_emails[j])

                        full_name = cc_full_names[j].split(' ') if len(cc_full_names) == len(cc_emails) else []
                        first_name = compile_name(full_name)['first']
                        last_name = compile_name(full_name)['last']

                        try:
                            company = re.search('@(.*)\.[com|ca|org|net]', cc_emails[j]).group(1)
                        except:
                            company = ''

                        output_worksheet['A' + str(current_row_in_results)].value = cc_emails[j]
                        output_worksheet['B' + str(current_row_in_results)].value = first_name
                        output_worksheet['C' + str(current_row_in_results)].value = last_name
                        output_worksheet['D' + str(current_row_in_results)].value = company

                        current_row_in_results += 1

    print("Completed contact compilation")
    return output_wb.save('compiled_contacts.xlsx')


def consolidate_raw_list(input_wb_name, output_wb_name):
    input_wb = open_wb(input_wb_name)
    output_wb = open_wb(output_wb_name)
    compile_contacts(input_wb, output_wb)


def compare_and_remove_duplicates(master_wb_name, comparison_wb_name):
    master_wb = open_wb(master_wb_name)
    comparison_wb = open_wb(comparison_wb_name)
    remove_duplicates(master_wb, comparison_wb)


if __name__ == '__main__':
    consolidate_raw_list('heather_contacts_17MAR2022.xlsx', 'compiled_contacts.xlsx')
    compare_and_remove_duplicates('josh_compiled_contacts_final.xlsx','compiled_contacts.xlsx')
    